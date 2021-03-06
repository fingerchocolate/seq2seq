#xlsxファイルを読み込み、test or train/original or simpleの計4つのtxtファイルを作成するコード
#だらだら書いてるから入出力ファイルはコード内で指定しとるし、全くdefってない状態
#データ数train:test = N : 1-N
from __future__ import unicode_literals

import argparse
import collections
import io
import re

import progressbar
import random
import openpyxl as px##xlsxファイル読み込み用
import MeCab##分かち書き用
mecab = MeCab.Tagger ("-Owakati")##
MIN = 3 #最低単語数(未満は足切り)
N = 100 #データ数test:train = 1: N-1

split_pattern = re.compile(r'[^\w\s]')

nochange = []#平易化前後で不変の文対のリスト[(平易化前,平易化後)]
change=[]#平易化されている文対のリスト
book = px.load_workbook('T15-2018.2.28.xlsx')#xlsxファイル読み込み
#book = px.load_workbook('mini.xlsx')#xlsxファイル読み込み
active_sheet = book.active##アクティブシート
n_lines = active_sheet.max_row - 1
bar = progressbar.ProgressBar()
reject_cell = []
for i in range(n_lines):##
    some = "B" + str(i + 2)##セルの位置#平易化前B平易化後C
    print(some)
    s = active_sheet[some].value ##指定したセルの要素
    s = mecab.parse(s) ##文を分かち書き
    words = [] #リストの宣言
    for word in s.strip().split():##
        words.extend(split_pattern.split(word))
    words = [w for w in words if w] #単語毎を要素としてリスト化した一文
    print(words)
    print(len(words)) #一文の単語数
    if len(words) < MIN : ############
        reject_cell.append(some)#単語数による足切り文(行)

for i in range(n_lines):
    some_b = "B" + str(i + 2)##セルの位置#平易化前B平易化後C
    if some_b not in reject_cell:##
        s_b = active_sheet[some_b].value ##指定したセルの要素
        sw_b = mecab.parse(s_b)
        words_b = [] #リストの宣言
        for word_b in sw_b.strip().split():##
            words_b.extend(split_pattern.split(word_b))
        words_b = [w for w in words_b if w] #単語毎を要素としてリスト化した一文
        print(len(words_b)) #一文の単語数
        some_c = "C" + str(i + 2)##セルの位置#平易化前B平易化後C
        s_c = active_sheet[some_c].value ##指定したセルの要素
        if   s_c == s_b:
            nochange.append((len(words_b),s_b, s_c))
        else:
            change.append((len(words_b),s_b,s_c))
print(nochange)
print('平易化前後で不変',len(nochange))
print('平易化処理なされてる',len(change))

nochange.sort() #一文の単語数順にソート
change.sort()
print(nochange)
print(change)
test = [] #testデータ = [(単語数, 平易化前, 平易化後), (......),.....]
train =[] #trainデータ
for i, line in enumerate(nochange): #test:train=1:N-1データ数に振り分けていく
    if i % N == 0: #文の数
        test.append(line)
    else:
        train.append(line)

for i, line in enumerate(change):
    if i % N == 0: #文の数
        test.append(line)
    else:
        train.append(line)

random.shuffle(test) #リスト内の要素をシャッフル
random.shuffle(train)

with io.open('orignal_test.txt', 'w', encoding= 'utf-8') as f:
    with io.open('simple_test.txt', 'w', encoding= 'utf-8') as f2:
        for (len, s_b, s_c) in test:
            f.write(s_b)
            f.write('\n')
            f2.write(s_c)
            f2.write('\n')

with io.open('original_train.txt', 'w', encoding= 'utf-8') as f:
    with io.open('simple_train.txt', 'w', encoding= 'utf-8') as f2:
        for (len, s_b, s_c) in train:
                f.write(s_b)
                f.write('\n')
                f2.write(s_c)
                f2.write('\n')
