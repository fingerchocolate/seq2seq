#txtファイルを読み込んで分かち書きしたのをtxtファイルに書き込む
#$ python3 preprocess.py (読み込みソース(平易化前)ファイル).txt (書き込み用ソースファイル).txt (読み込み用ターゲット(平易化後)ファイル).txt (書き込み用ターゲットファイル) \--vocab-file (書き込み用平易化前後合わせたボキャファイル).txt
#これをtrain, testデータそれぞれに対して実行する.(但し, オプションのボキャファイルの作成はtrainのみ)
#$ python3 preprocess.py original_test.txt original_test_prepro.txt simple_test.txt simple_test_prepro.txt
#$ python3 preprocess.py original_train.txt original_train_prepro.txt simple_train.txt simple_train_prepro.txt \--vocab-file vocab_train.txt
#chainer/seq2seqのtutorial的なの(英仏翻訳)/前処理wmt_preprocess.py
#https://github.com/chainer/chainer/blob/master/examples/seq2seq/wmt_preprocess.py
#原文は#(コメントアウト), 追加文は##　 途中からめんどくなってやってない

from __future__ import unicode_literals

import argparse
import collections
import io
import re

import progressbar

import openpyxl as px##xlsxファイル読み込み用
import MeCab##分かち書き用
mecab = MeCab.Tagger ("-Owakati")##

split_pattern = re.compile(r'[^\w\s]')
#split_pattern = re.compile(r'([.,!?"\':;)(])')
digit_pattern = re.compile(r'\d')


#def split_sentence(s, use_lower):
def split_sentence(s): #sentenceの分かち書き
#    if use_lower:
#        s = s.lower() #文字列をすべて小文字にする
#    s = s.replace('\u2019', "'")
    s = digit_pattern.sub('0', s)
    s = mecab.parse(s) ##文を分かち書き
    print(s)
    words = [] #リストの宣言
    for word in s.strip().split():##
        words.extend(split_pattern.split(word))
    words = [w for w in words if w]
    print(words)###############かくにん
    return words

def count_lines(path): #行数のカウント
    with io.open(path, encoding='utf-8', errors='ignore') as f:
        return sum([1 for _ in f])
#    book = px.load_workbook(path)##xlsxファイル読み込み
#    active_sheet = book.active##アクティブシート
#    return (active_sheet.max_row - 1)##





#def read_file(path, use_lower):
def read_file(path): #ファイルの読み込み
    n_lines = count_lines(path)
    bar = progressbar.ProgressBar()
    with io.open(path, encoding='utf-8', errors='ignore') as f:
#    book = px.load_workbook(path)##xlsxファイル読み込み
#    active_sheet = book.active##アクティブシート
#        for line in bar(f, max_value=n_lines): #つかわないけどここえらーでる
#            words = split_sentence(line, use_lower)
#            words = split_sentence(line)
        line = f.readline()##+
        while line:##+
            words = split_sentence(line)
            line = f.readline()##+
#    for i in range(n_lines):##
#        some = "C" + str(i + 2)##セルの位置#平易化前B平易化後C
#        sentence = active_sheet[some].value ##指定したセルの要素
#        words = split_sentence(sentence)##
#        i += 1##
            yield words


#def proc_dataset(
#        path, outpath, vocab_path=None, vocab_size=None, use_lower=False):
def proc_dataset(
        path_souce, outpath_souce, path_target, outpath_target, vocab_path=None, vocab_size=None):
    token_count = 0
    counts = collections.Counter()
    with io.open(outpath_souce, 'w', encoding='utf-8') as f:
        for words_souce in read_file(path_souce):
            line_souce = ' '.join(words_souce)
            print(len(words_souce))#########################
            f.write(line_souce)
            f.write('\n')
            #######
            #len = len(words)#####
            #print(len)###########
            if vocab_path:
                for word in words_souce:
                    counts[word] += 1
            token_count += len(words_souce)
    print('number of souce tokens: %d' % token_count)

    with io.open(outpath_target, 'w', encoding='utf-8') as f:
        for words_target in read_file(path_target):
            line_target = ' '.join(words_target)
            print(len(words_target))#########################
            f.write(line_target)
            f.write('\n')
            #######
            #len = len(words)#####
            #print(len)###########
            if vocab_path:
                for word in words_target:
                    counts[word] += 1

    if vocab_path and vocab_size:
        vocab = [word for (word, _) in counts.most_common(vocab_size)]
        with io.open(vocab_path, 'w', encoding='utf-8') as f:
            for word in vocab:
                f.write(word)
                f.write('\n')



def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'INPUT_SOUCE', help='input souce sentence data')##
    parser.add_argument(
        'OUTPUT_SOUCE', help='output souce sentence data')##
    parser.add_argument(
        'INPUT_TARGET', help='input target sentence data')##
    parser.add_argument(
        'OUTPUT_TARGET', help='output target sentence data')##
    parser.add_argument(
        '--vocab-file', help='vocabulary file to save')
    parser.add_argument(
        '--vocab-size', type=int, default=40000,
        help='size of vocabulary file')
#    parser.add_argument(
#        '--lower', action='store_true', help='use lower case')
    args = parser.parse_args()

    proc_dataset(
        args.INPUT_SOUCE, args.OUTPUT_SOUCE, args.INPUT_TARGET, args.OUTPUT_TARGET, vocab_path=args.vocab_file,vocab_size=args.vocab_size)
#        vocab_size=args.vocab_size, use_lower=args.lower)

if __name__ == '__main__':
	main()
