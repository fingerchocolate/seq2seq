#xlsxファイルを読み込み、
from __future__ import unicode_literals

import argparse
import collections
import io
import re

import progressbar

import openpyxl as px##xlsxファイル読み込み用
import MeCab##分かち書き用
mecab = MeCab.Tagger ("-Owakati")##

split_pattern = re.compile(r'[^\w\s]')
#split_pattern = re.compile(r'([.,!?"\':;)(])')
digit_pattern = re.compile(r'\d')


nochange = []#平易化前後で不変の文対のリスト[(平易化前,平易化後)]
change=[]#平易化されている文対のリスト
#book = px.load_workbook('T15-2018.2.28.xlsx')#xlsxファイル読み込み
book = px.load_workbook('sample.xlsx')#xlsxファイル読み込み
active_sheet = book.active##アクティブシート
n_lines = active_sheet.max_row - 1 
bar = progressbar.ProgressBar()
reject_line = []
for i in range(n_lines):##
    some = "B" + str(i + 2)##セルの位置#平易化前B平易化後C
    s = active_sheet[some].value ##指定したセルの要素
#    s = digit_pattern.sub('0', sentence)#数字を0に置換
    s = mecab.parse(s) ##文を分かち書き
    words = [] #リストの宣言
    for word in s.strip().split():##
        words.extend(split_pattern.split(word))
    words = [w for w in words if w] #単語毎を要素としてリスト化した一文
    i += 1##
    print(words)
    print(len(words)) #一文の単語数
    if len(words) < 3 : ############
        reject_line.append(i)#単語数による足切り文(行)

for i in range(n_lines):
    if i not in reject_line:##
        some_b = "B" + str(i + 2)##セルの位置#平易化前B平易化後C
        s_b = active_sheet[some_b].value ##指定したセルの要素
        #        s_c = digit_pattern.sub('0', sentence_c)
        sw_b = mecab.parse(s_b)
        words_b = [] #リストの宣言
        for word_b in sw_b.strip().split():##
            words_b.extend(split_pattern.split(word_b))
        words_b = [w for w in words_b if w] #単語毎を要素としてリスト化した一文
        print(len(words_b)) #一文の単語数
        some_c = "C" + str(i + 2)##セルの位置#平易化前B平易化後C
        s_c = active_sheet[some_c].value ##指定したセルの要素
#        s_b = digit_pattern.sub('0', sentence_b)
        if   s_c == s_b:
            nochange.append((len(words_b),s_b, s_c))
        else:
            change.append((len(words_b),s_b,s_c))
print(nochange)
print('不変',len(nochange))
print('変わってる',len(change))
ori_no = [s_b for (len, s_b,s_c) in nochange]
with io.open('gomi_ori_test.txt', 'w', encoding= 'utf-8') as f:
    with io.open('gomi_ori_train.txt', 'w', encoding= 'utf-8') as f2:
        for i, line in enumerate(ori_no):
            if i < 3: #文の数
                f.write(line)
                f.write('\n')
            else:
                f2.write(line)
                f2.write('\n')

sim_no = [s_c for (len, s_b,s_c) in nochange]
with io.open('gomi_sim_test.txt', 'w', encoding= 'utf-8') as f:
    with io.open('gomi_sim_train.txt', 'w', encoding= 'utf-8') as f2:
        for i, line in enumerate(sim_no):
            if i < 3: #文の数
                f.write(line)
                f.write('\n')
            else:
                f2.write(line)
                f2.write('\n')
                
ori_cha = [s_b for (len,s_b,s_c) in change]
with io.open('gomi_ori_test.txt', 'a', encoding= 'utf-8') as f3:
    with io.open('gomi_ori_train.txt', 'a', encoding= 'utf-8') as f4:
        for i, line in enumerate(ori_cha):
            if i < 3: #文の数
                f3.write(line)
                f3.write('\n')
            else:
                f4.write(line)
                f4.write('\n')

sim_cha = [s_c for (len, s_b,s_c) in change]
with io.open('gomi_sim_test.txt', 'a', encoding= 'utf-8') as f5:
    with io.open('gomi_sim_train.txt', 'a', encoding= 'utf-8') as f6:
        for i, line in enumerate(sim_cha):
            if i < 3: #文の数
                f5.write(line)
                f5.write('\n')
            else:
                f6.write(line)
                f6.write('\n')
