#文の単語数の最大・最小・平均もとめうる
#xlsxファイルを読み込んで分かち書きしたのをtextファイルに書き込む
#$ python3 preprocess.py (読み込みファイル).xlsx (書き込み用ファイル).txt
#chainer/seq2seqのtutorial的なの(英仏翻訳)/前処理wmt_preprocess.py
#https://github.com/chainer/chainer/blob/master/examples/seq2seq/wmt_preprocess.py
#原文は#(コメントアウト), 追加文は##

from __future__ import unicode_literals

import argparse
import collections
import io
import re
import numpy as np
import progressbar
import collections
import openpyxl as px##xlsxファイル読み込み用
import MeCab##分かち書き用
mecab = MeCab.Tagger ("-Owakati")##

split_pattern = re.compile(r'[^\w\s]')
#split_pattern = re.compile(r'([.,!?"\':;)(])')
digit_pattern = re.compile(r'\d')


#def split_sentence(s, use_lower):
def split_sentence(s): #sentenceの分かち書き
#    if use_lower:
#        s = s.lower() #文字列をすべて小文字にする
#    s = s.replace('\u2019', "'")
    s = digit_pattern.sub('0', s)
    s = mecab.parse(s) ##文を分かち書き
    #print(s)
    words = [] #リストの宣言
    for word in s.strip().split():##
        words.extend(split_pattern.split(word))
    words = [w for w in words if w]
    #print(words)###############かくにん
    return words

def count_lines(path): #行数のカウント
#    with io.open(path, encoding='utf-8', errors='ignore') as f:
    book = px.load_workbook(path)##xlsxファイル読み込み
    active_sheet = book.active##アクティブシート
    return (active_sheet.max_row - 1)##
#        return sum([1 for _ in f])


#def read_file(path, use_lower):
def read_file(path): #ファイルの読み込み
    n_lines = count_lines(path)
    bar = progressbar.ProgressBar()
#    with io.open(path, encoding='utf-8', errors='ignore') as f:
    book = px.load_workbook(path)##xlsxファイル読み込み
    active_sheet = book.active##アクティブシート
##         for line in bar(f, max_value=n_lines): #つかわないけどここえらーでる
#            words = split_sentence(line, use_lower)
#        line = f.readline()##+
#        while line:##+
#            words = split_sentence(line)
#            line = f.readline()##+
    for i in range(n_lines):##
        some = "C" + str(i + 2)##セルの位置#平易化前B平易化後C
        sentence = active_sheet[some].value ##指定したセルの要素
        words = split_sentence(sentence)##
        i += 1##
        yield words


#def proc_dataset(
#        path, outpath, vocab_path=None, vocab_size=None, use_lower=False):
def proc_dataset(
        path, outpath, vocab_path=None, vocab_size=None):
    token_count = 0
    counts = collections.Counter()
    lens = []##########
    #with io.open(outpath, 'w', encoding='utf-8') as f:
    for words in read_file(path):
        line = ' '.join(words)
        #print(len(words))##########################
        lens.append(len(words))##################
        if len(words)==1:######
            print(words)########
            #f.write(line)
            #f.write('\n')
            #######
        #if vocab_path:
           # for word in words:
                #counts[word] += 1
        #token_count += len(words)
    #print('number of tokens: %d' % token_count)

    #if vocab_path and vocab_size:
        #vocab = [word for (word, _) in counts.most_common(vocab_size)]
        #with io.open(vocab_path, 'w', encoding='utf-8') as f:
            #for word in vocab:
                #f.write(word)
                #f.write('\n')
    print("一文中に含まれる単語数")
    print("最大数:",np.amax(lens))
    print("最小数:",np.amin(lens))
    print("平均数:",np.mean(lens))
    c = collections.Counter(lens)
    print(c)

                

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'INPUT', help='input sentence data')
    parser.add_argument(
        'OUTPUT', help='output sentence data')
#    parser.add_argument(
#        '--vocab-file', help='vocabulary file to save')
#    parser.add_argument(
#        '--vocab-size', type=int, default=40000,
#        help='size of vocabulary file')
#    parser.add_argument(
#        '--lower', action='store_true', help='use lower case')
    args = parser.parse_args()

    proc_dataset(
        args.INPUT, args.OUTPUT,
#        vocab_path=args.vocab_file,
#        vocab_size=args.vocab_size
#        vocab_size=args.vocab_size, use_lower=args.lower)
        )
if __name__ == '__main__':
	main()
